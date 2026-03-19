"""
Data Exporter para guardar información de negocios en archivos CSV y Excel
"""

import pandas as pd
import os
from datetime import datetime
from typing import List, Dict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


class DataExporter:
    """Clase para exportar datos de negocios a diferentes formatos"""

    @staticmethod
    def to_csv(data: List[Dict[str, str]], filename: str = None, directory: str = "data") -> str:
        """
        Exporta datos de negocios a un archivo CSV

        Args:
            data: Lista de diccionarios con información de negocios
            filename: Nombre del archivo (si es None, se genera automáticamente)
            directory: Directorio donde guardar el archivo

        Returns:
            Ruta completa del archivo generado
        """
        if not data:
            raise ValueError("No hay datos para exportar")

        # Crear directorio si no existe
        os.makedirs(directory, exist_ok=True)

        # Generar nombre de archivo si no se proporciona
        if filename is None:
            timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
            filename = f"negocios_{timestamp}.csv"

        # Asegurar que termine en .csv
        if not filename.endswith('.csv'):
            filename += '.csv'

        # Ruta completa
        filepath = os.path.join(directory, filename)

        # Crear DataFrame
        df = pd.DataFrame(data)

        # Reordenar columnas en un orden lógico
        column_order = ['nombre', 'telefono', 'direccion', 'website', 'google_maps_url']
        # Mantener solo las columnas que existen
        available_columns = [col for col in column_order if col in df.columns]
        # Agregar columnas adicionales que no estén en el orden predefinido
        remaining_columns = [col for col in df.columns if col not in available_columns]
        df = df[available_columns + remaining_columns]

        # Guardar a CSV con codificación UTF-8 (para caracteres especiales como ñ, tildes)
        df.to_csv(filepath, index=False, encoding='utf-8-sig')  # utf-8-sig para Excel

        return filepath

    @staticmethod
    def print_summary(data: List[Dict[str, str]]):
        """
        Imprime un resumen de los datos extraídos

        Args:
            data: Lista de diccionarios con información de negocios
        """
        if not data:
            print("No hay datos para mostrar")
            return

        print("\n" + "="*60)
        print(f"📊 RESUMEN DE RESULTADOS")
        print("="*60)
        print(f"Total de negocios encontrados: {len(data)}")

        # Contar cuántos tienen cada tipo de información
        with_phone = sum(1 for b in data if b.get('telefono') and b['telefono'] != 'N/A')
        with_address = sum(1 for b in data if b.get('direccion') and b['direccion'] != 'N/A')
        with_website = sum(1 for b in data if b.get('website') and b['website'] != 'N/A')

        print(f"Con teléfono: {with_phone} ({with_phone/len(data)*100:.1f}%)")
        print(f"Con dirección: {with_address} ({with_address/len(data)*100:.1f}%)")
        print(f"Con sitio web: {with_website} ({with_website/len(data)*100:.1f}%)")
        print("="*60 + "\n")

    @staticmethod
    def to_excel(data: List[Dict], filename: str = None, directory: str = "data") -> str:
        """
        Exporta datos de negocios a un archivo Excel con formato y toda la información de contacto

        Args:
            data: Lista de diccionarios con información de negocios (debe incluir message_strategy)
            filename: Nombre del archivo (si es None, se genera automáticamente)
            directory: Directorio donde guardar el archivo

        Returns:
            Ruta completa del archivo generado
        """
        if not data:
            raise ValueError("No hay datos para exportar")

        # Crear directorio si no existe
        os.makedirs(directory, exist_ok=True)

        # Generar nombre de archivo si no se proporciona
        if filename is None:
            timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
            filename = f"negocios_{timestamp}.xlsx"

        # Asegurar que termine en .xlsx
        if not filename.endswith('.xlsx'):
            filename += '.xlsx'

        # Ruta completa
        filepath = os.path.join(directory, filename)

        # Crear workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Negocios"

        # Definir encabezados
        headers = [
            "Nombre",
            "Teléfono",
            "Dirección",
            "Website",
            "Google Maps",
            "Tipo Mensaje",
            "Mensaje WhatsApp",
            "Caso de Uso",
            "Timing Video Reservas",
            "Timing Video Estado",
            "Recomendaciones",
            "Seguimiento",
            "¿Contactado?"
        ]

        # Estilos para encabezados
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Escribir encabezados
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border

        # Escribir datos
        for row_num, business in enumerate(data, 2):
            strategy = business.get('message_strategy', {})

            # Preparar recomendaciones como texto con bullets
            recomendaciones = strategy.get('recomendaciones', [])
            recomendaciones_text = '\n'.join([f"• {rec}" for rec in recomendaciones])

            # Datos de la fila
            row_data = [
                business.get('nombre', 'N/A'),
                business.get('telefono', 'N/A'),
                business.get('direccion', 'N/A'),
                business.get('website', 'N/A'),
                business.get('google_maps_url', 'N/A'),
                f"MENSAJE {strategy.get('id', 'N/A')}: {strategy.get('nombre', 'N/A')}",
                strategy.get('mensaje_personalizado', 'N/A'),
                strategy.get('caso_uso', 'N/A'),
                strategy.get('timing_videos', {}).get('video_reservas', 'N/A'),
                strategy.get('timing_videos', {}).get('video_estado', 'N/A'),
                recomendaciones_text,
                strategy.get('seguimiento_personalizado', 'N/A'),
                ""  # Columna vacía para marcar manualmente
            ]

            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = value
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = border

                # Color de fondo para tipo de mensaje según la estrategia
                if col_num == 6:  # Columna "Tipo Mensaje"
                    color = strategy.get('color', '#FFFFFF').replace('#', '')
                    cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                    cell.font = Font(color="FFFFFF", bold=True)

        # Ajustar anchos de columna
        column_widths = {
            'A': 25,  # Nombre
            'B': 15,  # Teléfono
            'C': 30,  # Dirección
            'D': 20,  # Website
            'E': 15,  # Google Maps
            'F': 20,  # Tipo Mensaje
            'G': 60,  # Mensaje WhatsApp
            'H': 25,  # Caso de Uso
            'I': 40,  # Timing Video Reservas
            'J': 40,  # Timing Video Estado
            'K': 50,  # Recomendaciones
            'L': 50,  # Seguimiento
            'M': 12   # ¿Contactado?
        }

        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        # Congelar primera fila
        ws.freeze_panes = "A2"

        # Guardar
        wb.save(filepath)

        return filepath
